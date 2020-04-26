import openpyxl
from openpyxl.utils import get_column_letter
import networkx as nx
import matplotlib.pyplot as plt
"""
Running list of hard codes to be fixed:
1. .xlsx filename "CleanDataVersion.xlsx"
2. problem sheet name "Lesson1Analysis"
3. authorID in column C composed of all student IDs
4. heading of authorId column is "AuthorID"
5. student answers are directly to the right of the authorId column
6. There is only 1 "confirm" statement for this problem
"""

def display(graph):
    nx.draw(graph,with_labels=True)
    plt.show()

def main():
    # Open Excel sheet with student data
    students = set()
    wb = openpyxl.load_workbook('CleanDataCurrentVersion.xlsx',data_only=True)
    lesson1 = wb.get_sheet_by_name('Lesson1Analysis')
    authorID = lesson1['C']

    # make first pass through and get all student IDs
    flag = False
    for cell in authorID:
        if flag:
            if not (str(cell.value) in students) and not (cell.value == None):
                students.add(str(cell.value))
        else:
            if cell.value == "AuthorID":
                # set flag when we find start of column
                flag = True

    # initialize student dictionary with blank lists
    student_dict = {new_student : [] for new_student in students}

    # Parse into student dictionary
    for cell in authorID:
        student = str(cell.value)
        if student in students:
            col = get_column_letter(cell.column + 1)
            cell_num = "%s%d" % (col,cell.row)
            student_ans = str(lesson1[cell_num].value)
            student_dict[student].append(student_ans)
    
    # Traverse each student

    # Create graphs

    # Display graphs

    # Overlay and analyze graphs

    print("Waiting for user to terminate (Ctrl+c)")
    while True:
        pass

if __name__ == "__main__":
    main()
